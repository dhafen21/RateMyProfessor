import xlrd
from googlesearch import search
import openpyxl
import time

class Sheet:

    def __init__(self, p, sc_name):
        self.path = p
        self.school_name = sc_name
        self.search_array = []
        self.url_array = []
        self.no_url = []
        self.name_array = []
        self.check_list = []

    def get_urls(self):
        file_rd = xlrd.open_workbook(self.path)
        sheet = file_rd.sheet_by_index(0)
        index = 1
        for i in range(1, sheet.nrows):
        # for i in range(110,135):
            if i % 50 == 0:
                print("I need to wait a minute")
                time.sleep(10)
            # print(sheet.cell_value(i,0))
            a = sheet.cell_value(i,0).split(", ")
            string = "{} {} {} rate my professor ShowRatings".format(a[1], a[0], self.school_name)
            # string = "{} {} rate my professor".format(sheet.cell_value(i,0), self.school_name
            for j in search(string, 'co.in', num = 1, stop = 1, pause = 1.5):
                print(index)
                if "ratemyprofessors.com/ShowRatings" in j:
                  self.url_array.append(j)
                  self.name_array.append(sheet.cell_value(i,0))
                  break
                else:
                  # print("bad url: {}".format(j))
                  self.no_url.append(sheet.cell_value(i,0))
                  print("unable to find url for {}".format(sheet.cell_value(i,0)))
            index+=1
        return self.url_array



    def add_to_workbook(self, prof):
        file = openpyxl.load_workbook(self.path)
        for i in prof:
            n = i.name.split()
            name = "{}, {}".format(n[len(n)-1], n[0])
            newSheet = file.create_sheet(name)

            column = 1
            newSheet.cell(7,column).value = "Class Name"
            for j in range(8,len(i.data.class_name)+8):
                newSheet.cell(j, column).value = i.data.class_name[j - 8]

            column+=1
            newSheet.cell(7,column).value = "Quality"
            for j in range(8,len(i.data.class_name)+8):
                newSheet.cell(j, column).value = i.data.quality[j - 8]

            column+=1
            newSheet.cell(7,column).value = "Difficulty"
            for j in range(8,len(i.data.class_name)+8):
                newSheet.cell(j, column).value = i.data.difficulty[j - 8]

            column = 20
            newSheet.cell(7,column).value = "Comment Text"
            for j in range(8,len(i.data.class_name)+8):
                newSheet.cell(j, column).value = i.data.comments[j - 8]

            file.save(self.path)
        for i in range(0, len(prof)):
          n = prof[i].name.split()
          name = "{}, {}".format(n[len(n)-1], n[0])
          if self.name_array[i] != name:
            self.check_list.append("Excel: {}, found: {}".format(self.name_array[i], name))
